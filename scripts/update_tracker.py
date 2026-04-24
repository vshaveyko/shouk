#!/usr/bin/env python3
"""
Update bug tracker spreadsheet columns for a given bug ID.

Usage:
  python3 scripts/update_tracker.py SHK-001 --status "In Progress" --assigned-to "John"
  python3 scripts/update_tracker.py SHK-002 --date-resolved "2026-04-15" --resolution-notes "Fixed in v1.2"
  python3 scripts/update_tracker.py SHK-003 --additional-comments "Reproduced on mobile"

Writable columns:
  --status              Status
  --assigned-to         Assigned To
  --date-resolved       Date Resolved
  --additional-comments Additional Comments
  --resolution-notes    Resolution Notes
"""

import argparse
import io
import os
import sys

SERVICE_ACCOUNT_KEY = os.path.join(os.path.dirname(__file__), "sheets-bot-key.json")
SPREADSHEET_ID = "1lJOyELzWw3XD-KcW3Ll-Izfp5fUcwvo1"
SHEET_NAME = "Bugs"  # default; override with --sheet
WORKFLOWS_SHEET_NAME = "Human Verified Workflows"
ID_COLUMN = "ID"

WRITABLE_COLUMNS = {
    "status": "Status",
    "assigned_to": "Assigned To",
    "date_resolved": "Date Resolved",
    "additional_comments": "Additional Comments",
    "resolution_notes": "Resolution Notes",
}


def get_drive_service():
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError:
        print("Missing dependencies. Run: pip3 install google-api-python-client google-auth")
        sys.exit(1)

    if not os.path.exists(SERVICE_ACCOUNT_KEY):
        print(f"Service account key not found at: {SERVICE_ACCOUNT_KEY}")
        print("Copy the key file there or update SERVICE_ACCOUNT_KEY in this script.")
        sys.exit(1)

    creds = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_KEY,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    return build("drive", "v3", credentials=creds)


def download_workbook(drive):
    from googleapiclient.http import MediaIoBaseDownload
    import openpyxl

    request = drive.files().get_media(fileId=SPREADSHEET_ID)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return openpyxl.load_workbook(buf)


def upload_workbook(drive, wb):
    from googleapiclient.http import MediaIoBaseUpload

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    drive.files().update(fileId=SPREADSHEET_ID, media_body=media).execute()


def update_row(bug_id: str, updates: dict, sheet_name: str = SHEET_NAME):
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Missing dependency. Run: pip3 install openpyxl")
        sys.exit(1)

    drive = get_drive_service()

    print(f"Downloading spreadsheet...")
    wb = download_workbook(drive)
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]

    # Resolve column indices
    try:
        id_col = headers.index(ID_COLUMN) + 1
    except ValueError:
        print(f"Column '{ID_COLUMN}' not found in sheet '{SHEET_NAME}'.")
        sys.exit(1)

    col_indices = {}
    for key, col_name in WRITABLE_COLUMNS.items():
        if key in updates:
            try:
                col_indices[key] = headers.index(col_name) + 1
            except ValueError:
                print(f"Warning: column '{col_name}' not found in sheet, skipping.")

    # Find the row with the given bug ID
    target_row = None
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col - 1].value) == bug_id:
            target_row = row[0].row
            break

    if target_row is None:
        print(f"Bug ID '{bug_id}' not found in the tracker.")
        sys.exit(1)

    # Apply updates
    applied = []
    for key, col_idx in col_indices.items():
        col_name = WRITABLE_COLUMNS[key]
        value = updates[key]
        ws.cell(row=target_row, column=col_idx).value = value
        applied.append(f"  {col_name} = {value!r}")

    if not applied:
        print("Nothing to update.")
        return

    print(f"Uploading changes for {bug_id} (row {target_row}):")
    for line in applied:
        print(line)

    upload_workbook(drive, wb)
    print("Done.")


def main():
    parser = argparse.ArgumentParser(
        description="Update bug tracker columns for a given bug ID.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("bug_id", help="Bug ID, e.g. SHK-001 or A1.4")
    parser.add_argument("--sheet", default=SHEET_NAME,
                        help=f"Sheet name (default: '{SHEET_NAME}', workflows: '{WORKFLOWS_SHEET_NAME}')")
    parser.add_argument("--status", help="Status value")
    parser.add_argument("--assigned-to", dest="assigned_to", help="Assigned To value")
    parser.add_argument("--date-resolved", dest="date_resolved", help="Date Resolved value")
    parser.add_argument("--additional-comments", dest="additional_comments", help="Additional Comments value")
    parser.add_argument("--resolution-notes", dest="resolution_notes", help="Resolution Notes value")

    args = parser.parse_args()

    updates = {
        key: getattr(args, key)
        for key in WRITABLE_COLUMNS
        if getattr(args, key) is not None
    }

    if not updates:
        parser.print_help()
        sys.exit(1)

    update_row(args.bug_id, updates, sheet_name=args.sheet)


if __name__ == "__main__":
    main()
